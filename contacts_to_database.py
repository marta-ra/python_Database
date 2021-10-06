import sqlite3
import os
from openpyxl import load_workbook
import re
import datetime as DT


def cell_value(col):
    global worksheet
    value = worksheet.cell(row=rownum, column=col).value
    return value

sheets_dict = {'M - O': (1, 2, 3, 4, 5, 6, 'M', 'O', 23), 'M - E': (1, 3, 4, 5, 6, 2, 'M', 'E', 8),
               'M - B': (1, 2, 3, 4, 5, 6, 'M', 'B', 9), 'I - L - Q': (1, 2, 3, 5, 4, 6, 'I', 'L - Q', 0),
               'R E - S': (1, 2, 3, 5, 4, 6, 'R E', 'S', 0), 'M - R': (1, 2, 3, 4, 5, 6, 'M', 'R', 11),
               'I - L - C': (1, 3, 'N/A', 4, 5, 2, 'I', 'L - C', 0)}

# Connect to DB:
connection = sqlite3.connect('db_Contacts.db')
cur = connection.cursor()

# Read xlsx:
path = r''
workbook = load_workbook(os.path.join(path, "MYSheet.xlsx"))

for sheet_name in sheets_dict.keys():
    for sheet in workbook:
        if sheet.title == sheet_name:
            workbook[sheet.title].views.sheetView[0].tabSelected = True
        else:
            workbook[sheet.title].views.sheetView[0].tabSelected = False
    workbook.active = workbook[sheet_name]
    worksheet = workbook.active
    # To dynamically find the last filled row
    max_row_not_empty = worksheet.max_row

    for rownum in range(2, max_row_not_empty + 1):
        print(rownum)
        # search in the file for the required values:
        date_created_value = cell_value(sheets_dict.get(sheet_name)[0])
        if date_created_value is None:
            date_created = 'null'
        else:
            data_find = (re.search(r'\d*\.\d*\.\d*', str(date_created_value))).group(0)
            month = ((re.search(r'\.\d*\.', data_find)).group(0)).replace('.', '')
            day = ((re.search(r'\b\d*.', data_find)).group(0)).replace('.', '')
            year = (re.search(r'\d*$', data_find)).group(0)
            date_created = f'{year}-{month}-{day}'
            print(date_created)
            # for translation in unix time:
            date_created = DT.datetime.strptime(date_created, '%Y-%m-%d')
            date_created = int(date_created.timestamp())
        first_name = cell_value(sheets_dict.get(sheet_name)[1])
        if first_name is None:
            first_name = 'null'
        if sheet_name == 'I - L - C' or sheet_name == 'R E - S':
            last_name = 'null'
        else:
            last_name = cell_value(sheets_dict.get(sheet_name)[2])
            if last_name is None:
                last_name = 'null'
        phone = cell_value(sheets_dict.get(sheet_name)[3])
        if phone is None:
            phone = 'null'
        email = cell_value(sheets_dict.get(sheet_name)[4])
        if email is None:
            email = 'null'
        province = cell_value(sheets_dict.get(sheet_name)[5])
        if province is None:
            province = 'null'
        niche = sheets_dict.get(sheet_name)[6]
        l_type = sheets_dict.get(sheet_name)[7]
        if sheet_name in ['I - L - Q', 'R E - S', 'I - L - C']:
            c_score = 0
        else:
            c_score_find = cell_value(sheets_dict.get(sheet_name)[8])
            if c_score_find is None:
                c_score = 0
            else:
                try:
                    c_score = int((re.search(r'\d\d\d', str(c_score_find))).group(0))
                except:
                    credit_score = 0
        print(first_name)
        # adding a new contact to the database(with SQL):
        if first_name != 'Test' and first_name != 'test':
            values = (date_created, first_name, last_name, phone, email, province, niche, l_type, c_score)
            print(values)
            sql_str = f'INSERT INTO contacts(date_created, first_name, last_name, phone, email, province, niche, l_type, c_score) VALUES {values};'
            cur.execute(sql_str)
            connection.commit()

cur.close()
